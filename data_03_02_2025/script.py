from openpyxl import load_workbook

workbook = load_workbook('workspace.xlsx')

print(workbook.sheetnames)

planilha = workbook['cadastros']

print(planilha['C4'].value)

planilha.cell(row=4, column=3).value = '000.000.000-00'

workbook.save('workspace.xlsx')